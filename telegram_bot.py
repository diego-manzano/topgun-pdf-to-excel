#!/usr/bin/env python3
"""
Telegram Bot for Petron Settlement Report Extraction
Receives images or PDFs, extracts data with Gemini, returns Excel
"""

import os
import json
import logging
from pathlib import Path
from datetime import datetime
from io import BytesIO

# Load environment variables from .env file
try:
    from dotenv import load_dotenv
    load_dotenv()  # Load .env file if it exists
except ImportError:
    # python-dotenv not installed, will use system environment variables
    pass

# Telegram bot imports
try:
    from telegram import Update
    from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
except ImportError:
    print("Installing python-telegram-bot...")
    os.system("pip install python-telegram-bot --break-system-packages -q")
    from telegram import Update
    from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

# Gemini imports
try:
    import google.generativeai as genai
except ImportError:
    print("Installing google-generativeai...")
    os.system("pip install google-generativeai --break-system-packages -q")
    import google.generativeai as genai

# Excel imports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl --break-system-packages -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Gemini extraction prompt
EXTRACTION_PROMPT = """
Extract all transaction data from this Petron Merchant Settlement Report.
The report may be an image or PDF with multiple pages. Focus on the transaction table.

Return as JSON with this exact structure:
{
  "header": {
    "customer_number": "",
    "business_location_id": "",
    "business_location_name": "",
    "date_from": "",
    "date_to": "",
    "reimbursement_batch": ""
  },
  "transactions": [
    {
      "terminal_id": "",
      "host_batch_id": "",
      "ids": "",
      "settle_date": "",
      "no_of_txn": 0,
      "gross_amount": 0.00,
      "ewt": 0.00,
      "net_amount": 0.00,
      "description": ""
    }
  ],
  "totals": {
    "gross_amount": 0.00,
    "ewt": 0.00,
    "net_amount": 0.00
  }
}

Important:
- Extract ALL transaction rows from the table
- Parse numbers correctly (remove commas from amounts)
- Keep dates in their original format
- Only return valid JSON, no markdown code blocks or extra text
"""

class GeminiService:
    """Service for extracting data using Gemini Vision API"""
    
    def __init__(self, api_key: str):
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel('gemini-2.5-flash')
        logger.info("Gemini service initialized with gemini-2.5-flash")
    
    async def extract_from_bytes(self, file_bytes: bytes, mime_type: str) -> dict:
        """Extract data from image or PDF bytes"""
        logger.info(f"Extracting data from {mime_type}, size: {len(file_bytes)} bytes")
        
        try:
            response = self.model.generate_content([
                EXTRACTION_PROMPT,
                {
                    "mime_type": mime_type,
                    "data": file_bytes
                }
            ])
            
            # Parse JSON response
            response_text = response.text.strip()
            
            # Remove markdown code blocks if present
            if response_text.startswith('```json'):
                response_text = response_text[7:]
            if response_text.startswith('```'):
                response_text = response_text[3:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]
            response_text = response_text.strip()
            
            data = json.loads(response_text)
            logger.info(f"Successfully extracted {len(data.get('transactions', []))} transactions")
            
            return data
            
        except json.JSONDecodeError as e:
            logger.error(f"JSON parsing error: {e}")
            raise ValueError("Failed to parse Gemini response as JSON")
        except Exception as e:
            logger.error(f"Extraction error: {e}")
            raise

class ExcelService:
    """Service for generating Excel files"""
    
    @staticmethod
    def generate_report(data: dict) -> bytes:
        """Generate formatted Excel file from extracted data"""
        logger.info("Generating Excel report")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Settlement Report"
        
        # Define styles
        title_font = Font(name='Arial', size=14, bold=True)
        header_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        currency_format = '#,##0.00'
        
        # Border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header fill
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        
        # === TITLE SECTION ===
        ws['A1'] = "Merchant Settlement Report"
        ws['A1'].font = title_font
        ws.merge_cells('A1:I1')
        
        # === HEADER INFO SECTION ===
        current_row = 3
        
        # Left side info
        ws[f'A{current_row}'] = "Customer Number:"
        ws[f'B{current_row}'] = data['header']['customer_number']
        ws[f'A{current_row}'].font = header_font
        
        current_row += 1
        ws[f'A{current_row}'] = "Business Location:"
        ws[f'B{current_row}'] = data['header']['business_location_id']
        ws[f'C{current_row}'] = data['header']['business_location_name']
        ws[f'A{current_row}'].font = header_font
        
        # Right side info
        ws['G3'] = "From:"
        ws['H3'] = data['header']['date_from']
        ws['G3'].font = header_font
        
        ws['G4'] = "To:"
        ws['H4'] = data['header']['date_to']
        ws['G4'].font = header_font
        
        ws['G5'] = "Reimbursement Batch:"
        ws['H5'] = data['header']['reimbursement_batch']
        ws['G5'].font = header_font
        
        # === TABLE HEADERS ===
        table_start_row = 7
        headers = [
            "Terminal ID", "Host Batch ID", "Ids", "Settle Date",
            "No Of Txn", "Transaction\nGross Amount", "EWT",
            "Transaction\nNet Amount", "Description"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # === TRANSACTION ROWS ===
        current_row = table_start_row + 1
        
        for txn in data['transactions']:
            ws.cell(row=current_row, column=1, value=txn['terminal_id']).border = thin_border
            ws.cell(row=current_row, column=2, value=txn['host_batch_id']).border = thin_border
            ws.cell(row=current_row, column=3, value=txn['ids']).border = thin_border
            ws.cell(row=current_row, column=4, value=txn['settle_date']).border = thin_border
            
            txn_cell = ws.cell(row=current_row, column=5, value=txn['no_of_txn'])
            txn_cell.border = thin_border
            txn_cell.alignment = Alignment(horizontal='center')
            
            gross_cell = ws.cell(row=current_row, column=6, value=txn['gross_amount'])
            gross_cell.number_format = currency_format
            gross_cell.border = thin_border
            
            ewt_cell = ws.cell(row=current_row, column=7, value=txn['ewt'])
            ewt_cell.number_format = currency_format
            ewt_cell.border = thin_border
            
            net_cell = ws.cell(row=current_row, column=8, value=txn['net_amount'])
            net_cell.number_format = currency_format
            net_cell.border = thin_border
            
            desc_cell = ws.cell(row=current_row, column=9, value=txn['description'])
            desc_cell.border = thin_border
            
            current_row += 1
        
        # === TOTALS ROW ===
        totals_row = current_row
        ws.cell(row=totals_row, column=5, value="Total:").font = header_font
        
        total_gross = ws.cell(row=totals_row, column=6, value=data['totals']['gross_amount'])
        total_gross.number_format = currency_format
        total_gross.font = header_font
        total_gross.border = thin_border
        
        total_ewt = ws.cell(row=totals_row, column=7, value=data['totals']['ewt'])
        total_ewt.number_format = currency_format
        total_ewt.font = header_font
        total_ewt.border = thin_border
        
        total_net = ws.cell(row=totals_row, column=8, value=data['totals']['net_amount'])
        total_net.number_format = currency_format
        total_net.font = header_font
        total_net.border = thin_border
        
        # === COLUMN WIDTHS ===
        column_widths = {
            'A': 12, 'B': 13, 'C': 10, 'D': 20, 'E': 10,
            'F': 16, 'G': 10, 'H': 16, 'I': 40,
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        ws.row_dimensions[table_start_row].height = 30
        
        # === SAVE TO BYTES ===
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Excel generated: {len(data['transactions'])} transactions")
        return buffer.getvalue()

# Initialize services
gemini_service = None
excel_service = ExcelService()

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Send welcome message"""
    await update.message.reply_text(
        "👋 **Welcome to the Settlement Report Bot!**\n\n"
        "Send me a Petron Merchant Settlement Report as a **PDF document**.\n\n"
        "I'll extract the data and send you a formatted Excel file.\n\n"
        "**Commands:**\n"
        "/start - Show this message\n"
        "/help - Usage instructions",
        parse_mode='Markdown'
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Send help message"""
    await update.message.reply_text(
        "📖 **How to use:**\n\n"
        "1. Get a PDF of your Petron settlement report\n"
        "2. Send it to this bot as a document\n"
        "3. Wait 5-10 seconds for processing\n"
        "4. Download your Excel file!\n\n"
        "**Supported format:**\n"
        "• PDF documents only\n\n"
        "**Tips:**\n"
        "• Ensure the PDF is clear and readable\n"
        "• All transaction rows should be visible\n"
        "• Works best with standard Petron reports",
        parse_mode='Markdown'
    )

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Reject photo messages and ask for PDF"""
    await update.message.reply_text(
        "❌ **Photos are not supported**\n\n"
        "Please send your settlement report as a **PDF document**.\n\n"
        "To send a PDF:\n"
        "1. Click the paperclip/attachment icon\n"
        "2. Select 'Document' or 'File'\n"
        "3. Choose your PDF file\n"
        "4. Send it to me!",
        parse_mode='Markdown'
    )

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle document messages (PDFs only)"""
    try:
        document = update.message.document
        logger.info(f"Received document: {document.file_name} from user {update.effective_user.id}")
        
        # Check if it's a PDF
        mime_type = document.mime_type
        if mime_type != 'application/pdf':
            await update.message.reply_text(
                "❌ **Only PDF files are supported**\n\n"
                f"You sent: {document.file_name}\n"
                f"Type: {mime_type}\n\n"
                "Please send your settlement report as a **PDF document**.",
                parse_mode='Markdown'
            )
            return
        
        # Download PDF
        file = await document.get_file()
        file_bytes = await file.download_as_bytearray()
        
        await process_file(update, bytes(file_bytes), mime_type)
        
    except Exception as e:
        logger.error(f"Error handling document: {e}")
        await update.message.reply_text(
            f"❌ Error processing PDF: {str(e)}\n\n"
            "Please ensure the file is a valid Petron settlement report."
        )

async def process_file(update: Update, file_bytes: bytes, mime_type: str):
    """Process PDF file and return Excel"""
    # Send processing message
    processing_msg = await update.message.reply_text("🔄 Processing your report...")
    
    try:
        # Extract data using Gemini
        logger.info("Extracting data with Gemini...")
        data = await gemini_service.extract_from_bytes(file_bytes, mime_type)
        
        # Generate Excel
        logger.info("Generating Excel file...")
        excel_bytes = excel_service.generate_report(data)
        
        # Delete processing message
        await processing_msg.delete()
        
        # Send Excel file
        filename = f"settlement_report_{data['header']['reimbursement_batch']}.xlsx"
        
        await update.message.reply_document(
            document=BytesIO(excel_bytes),
            filename=filename,
            caption=(
                f"✅ **Report extracted successfully!**\n\n"
                f"📊 **{len(data['transactions'])} transactions**\n"
                f"💰 **Total Net Amount:** ₱{data['totals']['net_amount']:,.2f}\n"
                f"📅 **Period:** {data['header']['date_from']} - {data['header']['date_to']}\n"
                f"🔢 **Batch:** {data['header']['reimbursement_batch']}"
            ),
            parse_mode='Markdown'
        )
        
        logger.info(f"Successfully processed report for batch {data['header']['reimbursement_batch']}")
        
    except ValueError as e:
        await processing_msg.edit_text(
            f"❌ **Extraction failed:** {str(e)}\n\n"
            "The report format may not be recognized. Please ensure it's a valid Petron settlement report."
        )
    except Exception as e:
        logger.error(f"Processing error: {e}", exc_info=True)
        await processing_msg.edit_text(
            f"❌ **Error:** {str(e)}\n\n"
            "Please try again or contact support if the issue persists."
        )

def main():
    """Start the bot"""
    # Check environment variables
    bot_token = os.getenv('TELEGRAM_BOT_TOKEN')
    gemini_api_key = os.getenv('GEMINI_API_KEY')
    webhook_url = os.getenv('WEBHOOK_URL')  # Optional: for webhook mode
    port = int(os.getenv('PORT', '8080'))  # Port for webhook
    
    if not bot_token:
        logger.error("TELEGRAM_BOT_TOKEN environment variable not set!")
        print("\n❌ Missing TELEGRAM_BOT_TOKEN")
        print("\nTo create a bot:")
        print("1. Open Telegram and search for @BotFather")
        print("2. Send /newbot and follow instructions")
        print("3. Copy the token and set it: export TELEGRAM_BOT_TOKEN='your-token'")
        return
    
    if not gemini_api_key:
        logger.error("GEMINI_API_KEY environment variable not set!")
        print("\n❌ Missing GEMINI_API_KEY")
        print("\nTo get an API key:")
        print("1. Go to https://aistudio.google.com/")
        print("2. Click 'Get API Key'")
        print("3. Copy the key and set it: export GEMINI_API_KEY='your-key'")
        return
    
    # Initialize Gemini service
    global gemini_service
    gemini_service = GeminiService(gemini_api_key)
    
    # Create application
    logger.info("Starting Telegram bot...")
    app = Application.builder().token(bot_token).build()
    
    # Add handlers
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    
    # Determine mode: Webhook or Polling
    if webhook_url:
        # Webhook mode - for production deployment
        logger.info(f"🌐 Starting in WEBHOOK mode on port {port}")
        logger.info(f"   Webhook URL: {webhook_url}")
        print("\n" + "="*60)
        print("🌐 Settlement Report Bot - WEBHOOK MODE")
        print("="*60)
        print(f"\nWebhook URL: {webhook_url}")
        print(f"Port: {port}")
        print("\nBot will only run when receiving messages.")
        print("Zero resource usage when idle!")
        print("="*60 + "\n")
        
        app.run_webhook(
            listen="0.0.0.0",
            port=port,
            url_path=bot_token,
            webhook_url=f"{webhook_url}/{bot_token}"
        )
    else:
        # Polling mode - for local development
        logger.info("🔄 Starting in POLLING mode")
        print("\n" + "="*60)
        print("🤖 Settlement Report Bot - POLLING MODE")
        print("="*60)
        print("\nRunning locally with polling (good for development)")
        print("Waiting for messages...")
        print("\nTo use webhook mode (for production):")
        print("  Set WEBHOOK_URL environment variable")
        print("\nPress Ctrl+C to stop the bot")
        print("="*60 + "\n")
        
        app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()