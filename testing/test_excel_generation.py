#!/usr/bin/env python3
"""
Test Excel generation for Petron Settlement Reports
Takes extracted JSON data and creates a formatted Excel file
"""

import json
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import os
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

# Sample extracted data (from your successful test)
SAMPLE_DATA = {
  "header": {
    "customer_number": "1049850",
    "business_location_id": "100000040277201",
    "business_location_name": "Top Gun 747 Corporation",
    "date_from": "01 Nov 2025",
    "date_to": "03 Nov 2025",
    "reimbursement_batch": "5216"
  },
  "transactions": [
    {
      "terminal_id": "20020788",
      "host_batch_id": "28916273",
      "ids": "13398430",
      "settle_date": "11/01/2025 1:58PM",
      "no_of_txn": 3,
      "gross_amount": 5757.21,
      "ewt": 51.41,
      "net_amount": 5705.80,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "20020788",
      "host_batch_id": "28916604",
      "ids": "13398761",
      "settle_date": "11/01/2025 10:01PM",
      "no_of_txn": 2,
      "gross_amount": 997.64,
      "ewt": 8.91,
      "net_amount": 988.73,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28916857",
      "ids": "13399842",
      "settle_date": "11/01/2025 12:00AM",
      "no_of_txn": 7,
      "gross_amount": 8145.41,
      "ewt": 72.73,
      "net_amount": 8072.68,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28917136",
      "ids": "13399841",
      "settle_date": "11/01/2025 12:00AM",
      "no_of_txn": 15,
      "gross_amount": 17738.45,
      "ewt": 158.39,
      "net_amount": 17580.06,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28918111",
      "ids": "13399840",
      "settle_date": "11/01/2025 12:00AM",
      "no_of_txn": 3,
      "gross_amount": 4296.11,
      "ewt": 38.36,
      "net_amount": 4257.75,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28918115",
      "ids": "13399839",
      "settle_date": "11/01/2025 12:00AM",
      "no_of_txn": 3,
      "gross_amount": 3092.90,
      "ewt": 27.62,
      "net_amount": 3065.28,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28919726",
      "ids": "13402614",
      "settle_date": "11/02/2025 12:00AM",
      "no_of_txn": 1,
      "gross_amount": 1804.76,
      "ewt": 16.11,
      "net_amount": 1788.65,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28920099",
      "ids": "13402613",
      "settle_date": "11/02/2025 12:00AM",
      "no_of_txn": 34,
      "gross_amount": 42816.21,
      "ewt": 382.29,
      "net_amount": 42433.92,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28920933",
      "ids": "13402615",
      "settle_date": "11/02/2025 12:00AM",
      "no_of_txn": 13,
      "gross_amount": 13706.08,
      "ewt": 122.38,
      "net_amount": 13583.70,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "20020788",
      "host_batch_id": "28921640",
      "ids": "13403797",
      "settle_date": "11/03/2025 5:58AM",
      "no_of_txn": 1,
      "gross_amount": 2685.42,
      "ewt": 23.98,
      "net_amount": 2661.44,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "20020788",
      "host_batch_id": "28921956",
      "ids": "13404113",
      "settle_date": "11/03/2025 2:05PM",
      "no_of_txn": 3,
      "gross_amount": 5376.84,
      "ewt": 48.00,
      "net_amount": 5328.84,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "20020788",
      "host_batch_id": "28922369",
      "ids": "13404526",
      "settle_date": "11/03/2025 10:00PM",
      "no_of_txn": 2,
      "gross_amount": 2002.03,
      "ewt": 17.87,
      "net_amount": 1984.16,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28923152",
      "ids": "13405737",
      "settle_date": "11/03/2025 12:00AM",
      "no_of_txn": 31,
      "gross_amount": 38742.75,
      "ewt": 345.94,
      "net_amount": 38396.81,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28924539",
      "ids": "13405738",
      "settle_date": "11/03/2025 12:00AM",
      "no_of_txn": 4,
      "gross_amount": 7618.13,
      "ewt": 68.02,
      "net_amount": 7550.11,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28924666",
      "ids": "13405739",
      "settle_date": "11/03/2025 12:00AM",
      "no_of_txn": 2,
      "gross_amount": 1983.50,
      "ewt": 17.71,
      "net_amount": 1965.79,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28924737",
      "ids": "13405740",
      "settle_date": "11/03/2025 12:00AM",
      "no_of_txn": 19,
      "gross_amount": 21119.29,
      "ewt": 188.57,
      "net_amount": 20930.72,
      "description": "Default Fleet Transaction (Prod Level)"
    },
    {
      "terminal_id": "50035936",
      "host_batch_id": "28924760",
      "ids": "13405741",
      "settle_date": "11/03/2025 12:00AM",
      "no_of_txn": 2,
      "gross_amount": 687.55,
      "ewt": 6.14,
      "net_amount": 681.41,
      "description": "Default Fleet Transaction (Prod Level)"
    }
  ],
  "totals": {
    "gross_amount": 178570.28,
    "ewt": 1594.43,
    "net_amount": 176975.85
  }
}

def generate_excel(data: dict, output_path: str = None) -> str:
    """Generate formatted Excel file from extracted data"""
    
    print("📊 Generating Excel file...")
    print(f"   Transactions: {len(data['transactions'])}")
    print(f"   Customer: {data['header']['business_location_name']}")
    print()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Settlement Report"
    
    # Define styles
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    currency_format = '#,##0.00'
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header fill
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    # === TITLE SECTION ===
    ws['A1'] = "Merchant Settlement Report"
    ws['A1'].font = title_font
    ws.merge_cells('A1:I1')
    
    # === HEADER INFO SECTION ===
    current_row = 3
    
    # Left side info
    ws[f'A{current_row}'] = f"Customer Number:"
    ws[f'B{current_row}'] = data['header']['customer_number']
    ws[f'A{current_row}'].font = header_font
    
    current_row += 1
    ws[f'A{current_row}'] = f"Business Location:"
    ws[f'B{current_row}'] = data['header']['business_location_id']
    ws[f'C{current_row}'] = data['header']['business_location_name']
    ws[f'A{current_row}'].font = header_font
    
    # Right side info (dates and batch)
    ws['G3'] = "From:"
    ws['H3'] = data['header']['date_from']
    ws['G3'].font = header_font
    
    ws['G4'] = "To:"
    ws['H4'] = data['header']['date_to']
    ws['G4'].font = header_font
    
    ws['G5'] = "Reimbursement Batch:"
    ws['H5'] = data['header']['reimbursement_batch']
    ws['G5'].font = header_font
    
    # === TABLE HEADERS ===
    table_start_row = 7
    headers = [
        "Terminal ID",
        "Host Batch ID",
        "Ids",
        "Settle Date",
        "No Of Txn",
        "Transaction\nGross Amount",
        "EWT",
        "Transaction\nNet Amount",
        "Description"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=table_start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # === TRANSACTION ROWS ===
    current_row = table_start_row + 1
    
    for txn in data['transactions']:
        ws.cell(row=current_row, column=1, value=txn['terminal_id']).border = thin_border
        ws.cell(row=current_row, column=2, value=txn['host_batch_id']).border = thin_border
        ws.cell(row=current_row, column=3, value=txn['ids']).border = thin_border
        ws.cell(row=current_row, column=4, value=txn['settle_date']).border = thin_border
        
        # No of transactions (center aligned)
        txn_cell = ws.cell(row=current_row, column=5, value=txn['no_of_txn'])
        txn_cell.border = thin_border
        txn_cell.alignment = Alignment(horizontal='center')
        
        # Currency amounts
        gross_cell = ws.cell(row=current_row, column=6, value=txn['gross_amount'])
        gross_cell.number_format = currency_format
        gross_cell.border = thin_border
        
        ewt_cell = ws.cell(row=current_row, column=7, value=txn['ewt'])
        ewt_cell.number_format = currency_format
        ewt_cell.border = thin_border
        
        net_cell = ws.cell(row=current_row, column=8, value=txn['net_amount'])
        net_cell.number_format = currency_format
        net_cell.border = thin_border
        
        desc_cell = ws.cell(row=current_row, column=9, value=txn['description'])
        desc_cell.border = thin_border
        
        current_row += 1
    
    # === TOTALS ROW ===
    totals_row = current_row
    
    ws.cell(row=totals_row, column=5, value="Total:").font = header_font
    
    total_gross = ws.cell(row=totals_row, column=6, value=data['totals']['gross_amount'])
    total_gross.number_format = currency_format
    total_gross.font = header_font
    total_gross.border = thin_border
    
    total_ewt = ws.cell(row=totals_row, column=7, value=data['totals']['ewt'])
    total_ewt.number_format = currency_format
    total_ewt.font = header_font
    total_ewt.border = thin_border
    
    total_net = ws.cell(row=totals_row, column=8, value=data['totals']['net_amount'])
    total_net.number_format = currency_format
    total_net.font = header_font
    total_net.border = thin_border
    
    # === COLUMN WIDTHS ===
    column_widths = {
        'A': 12,  # Terminal ID
        'B': 13,  # Host Batch ID
        'C': 10,  # Ids
        'D': 20,  # Settle Date
        'E': 10,  # No Of Txn
        'F': 16,  # Gross Amount
        'G': 10,  # EWT
        'H': 16,  # Net Amount
        'I': 40,  # Description
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row height for header
    ws.row_dimensions[table_start_row].height = 30
    
    # === SAVE FILE ===
    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = f"settlement_report_{data['header']['reimbursement_batch']}_{timestamp}.xlsx"
    
    wb.save(output_path)
    
    print(f"✅ Excel file generated successfully!")
    print(f"   File: {output_path}")
    print(f"   Size: {Path(output_path).stat().st_size / 1024:.1f} KB")
    
    return output_path

def test_excel_generation():
    """Test the Excel generation with sample data"""
    
    print("🧪 Excel Generation Test")
    print("=" * 80)
    print()
    
    # Check if we have extracted_data.json from previous test
    json_path = Path("extracted_data.json")
    
    if json_path.exists():
        print("📂 Found extracted_data.json from previous test")
        with open(json_path, 'r') as f:
            data = json.load(f)
    else:
        print("📂 Using sample data embedded in script")
        data = SAMPLE_DATA
    
    # Generate Excel
    output_file = generate_excel(data)
    
    print()
    print("=" * 80)
    print("✅ Test completed successfully!")
    print()
    print(f"Open the file: {output_file}")
    print()
    print("Validation checklist:")
    print("  ✓ Header information (Customer, Business, Dates, Batch)")
    print("  ✓ Table headers with proper formatting")
    print(f"  ✓ {len(data['transactions'])} transaction rows")
    print("  ✓ Currency formatting (amounts with commas)")
    print("  ✓ Totals row with bold formatting")
    print("  ✓ Column widths adjusted for readability")
    
    return output_file

if __name__ == "__main__":
    test_excel_generation()